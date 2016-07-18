import os 
from openpyxl import Workbook
from openpyxl import load_workbook



#
#d:\desktop\shd1-2016.03

	#	new_list.insert(cell.col_idx,cell.value)
	#ws.append(new_list)
#ws.append(ws2.iter_rows())

def filter_column(idx,columns):
	for deny in columns:
		if idx == deny:
			return False
	return True

def deny(name,denys):
	for deny in denys:
		if name.find(deny) != -1:
			return False
	return True

def concate_grp(paths,denyed_columns):
	allthree=[]
	for path in paths:
		wb2 = load_workbook(path)
		ws2 = wb2.active
		listoflists=[]
		for row in ws2.iter_rows():
			new_list=[]
			row_idx = 1
			for cell in row:
				if cell.row > 1:
					row_idx = cell.row
					if filter_column(cell.col_idx,denyed_columns) == True:
						new_list.append(cell.value)
				else: 
					break
			if row_idx> 1:
				listoflists.insert((row_idx - 2),new_list)
		allthree.append(listoflists)
	return concate_horiz(allthree)

def concate_horiz(ls):
	newls=[]
	for index,lists in enumerate(ls):
		if index == 0:
			newls = lists
		else:
			for i, row in enumerate(lists):
				newls[i] = newls[i] + row
	return newls

def concate_1st_grp(paths,denyed_columns):
	allthree=[]
	for path in paths:
		wb2 = load_workbook(path)
		ws2 = wb2.active
		listoflists=[]

		for row in ws2.iter_rows():
			new_list=[]
			for cell in row:
				row_idx = cell.row
				if filter_column(cell.col_idx,denyed_columns) == True:
					new_list.append(cell.value)
			listoflists.append(new_list)
		allthree.append(listoflists)
	return concate_horiz(allthree)


def init():
	rootdir = input('Enter directory path:')
	denys = input('Enter exclude keywords (comma separated):').split(',')
	denyed_columns = input('Remove columns (comma separated):').split(',')

	wb = Workbook()

	# grab the active worksheet
	ws = wb.active
	titled = False
	file_dirs={}
	dcount = 0
	for entry in os.scandir(rootdir):
		if entry.is_dir(follow_symlinks=False):
			dcount = dcount + 1
			print(str(dcount)+': '+entry.name)
			remove_prefix = entry.name[6:]
			end = remove_prefix.find('-')
			if(end==-1):
				key = int(remove_prefix[0:])
			else:
				key = int(remove_prefix[0:remove_prefix.find('-')])
			file_dirs[key] = entry.path

	for key in sorted(file_dirs):
			newpaths=[]
			print(file_dirs[key])
			for entry2 in os.scandir(file_dirs[key]):
				if 'xls' in entry2.path or 'xlsx' in entry2.path:
					if deny(entry2.name,denys) == True:
						newpaths.append(entry2.path)
			if titled == False:
				lists = concate_1st_grp(newpaths,denyed_columns)
				for lst in lists:
					ws.append(lst)
				titled = True
			else:
				lists = concate_grp(newpaths,denyed_columns)
				for lst in lists:
					ws.append(lst)

	# Save the file
	wb.save(rootdir+'/combined.xlsx')



init()